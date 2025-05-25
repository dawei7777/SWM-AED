import argparse
import logging
import torch
import torch.nn as nn
import torch.nn.functional as F
from torchvision import datasets, transforms
import numpy as np
# from mnist_net import mnist_net
from visualize_save import visualize_images,visualize_images_x
from utils.readData import read_dataset
from utils.ResNet import ResNet18
from PIL import Image
import matplotlib.pyplot as plt
import csv
import torchattacks
from YOAO import yoao
import openpyxl
from openpyxl import Workbook
import re
# from models import *
def get_args():
    parser = argparse.ArgumentParser()
    parser.add_argument('--batch-size', default=1, type=int)
    parser.add_argument('--data-dir', default='../mnist-data', type=str)
    parser.add_argument('--epochs', default=1, type=int)
    parser.add_argument('--attack', default='fgsm', type=str, choices=['none', 'pgd', 'fgsm','sgsm','fgsm(x)'])
    parser.add_argument('--epsilon', default=0.001, type=float)
    parser.add_argument('--alpha', default=0.375, type=float)
    parser.add_argument('--attack-iters', default=40, type=int)
    parser.add_argument('--fname', default='mnist_model_none', type=str)
    parser.add_argument('--seed', default=0, type=int)
    return parser.parse_args()

from PIL import Image
import torch
from torchvision import transforms



# 图像预处理
preprocess = transforms.Compose([
    transforms.Resize(32),
    transforms.ToTensor(),
    transforms.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
])


def apply_mask(image, mask_size, position, color=(0, 1, 0)):
    """
    在图像上应用一个指定颜色的mask遮挡。

    :param image: 输入图像张量，形状应为[1, C, H, W]。
    :param mask_size: 遮挡的大小，例如(3, 3)。
    :param position: 遮挡的起始位置，例如(12, 5)。
    :param color: 遮挡的颜色，RGB值，例如(0, 1, 0)表示绿色。
    :return: 应用遮挡后的图像张量。
    """
    # 创建一个指定颜色的mask张量
    # mask = torch.tensor(color, dtype=torch.float32).view(3, 1, 1).repeat(1, mask_size[0], mask_size[1])
    mask = torch.tensor(color, dtype=image.dtype).view(3, 1, 1).repeat(1, mask_size[0], mask_size[1])
    # 将mask放置到图像的指定位置
    image[:, :, position[0]:position[0] + mask_size[0], position[1]:position[1] + mask_size[1]] = mask
    return image


def main():
    args = get_args()

    np.random.seed(args.seed)
    torch.manual_seed(args.seed)
    torch.cuda.manual_seed(args.seed)

    device = 'cuda' if torch.cuda.is_available() else 'cpu'
    n_class = 10
    model = ResNet18()  # 得到预训练模型
    # model = VGG('VGG11', n_class)
    # model = ResNet50(n_class)  # 得到预训练模型
    model.conv1 = nn.Conv2d(in_channels=3, out_channels=64, kernel_size=3, stride=1, padding=1, bias=False)
    model.fc = torch.nn.Linear(512, n_class)  # 将最后的全连接层修改
    # 载入权重
    model.load_state_dict(torch.load('checkpoint/resnet18_cifar10.pt'))
    # model.load_state_dict(torch.load('checkpoint/vgg11_cifar10_low.pt'))
    # model.load_state_dict(torch.load('checkpoint/resnet18_cifar10_low.pt'))
    # model.load_state_dict(torch.load('checkpoint/vgg11_cifar10_low-80%.pt'))
    model = model.to(device)
    model.eval()  # 验证模型

    criterion = nn.CrossEntropyLoss()

    batch_size = 1
    train_loader, valid_loader, test_loader = read_dataset(batch_size=batch_size, pic_path='dataset')
    attacks = [
        torchattacks.JSMA(model, theta=1.0, gamma=0.1),
        torchattacks.PGD(model, eps=8 / 255, alpha=1 / 255, steps=10, random_start=True),
        torchattacks.DeepFool(model, steps=10, overshoot=0.02),
        torchattacks.FGSM(model, eps=16 / 255),
        torchattacks.BIM(model, eps=8 / 255, alpha=2 / 255, steps=10),
        torchattacks.FFGSM(model, eps=8 / 255, alpha=10 / 255),
        torchattacks.APGD(model, norm='Linf', eps=8 / 255, steps=10, n_restarts=1, seed=0, loss='ce', eot_iter=1,
                          rho=.75, verbose=False),
        torchattacks.OnePixel(model, pixels=10, steps=10, popsize=10, inf_batch=128),
        torchattacks.Pixle(model, x_dimensions=(0.1, 0.2), restarts=10, max_iterations=10),
        torchattacks.PIFGSMPP(model, num_iter_set=10)
        # _, _, _, _, perturb_image = yoao(input_image.detach(), y, model, 10,overshoot=0.02, max_iter=1)
    ]



    # thresholds = np.linspace(0.053, 1, 20)
    thresholds = np.linspace(0.003, 3.32, 332)


    for attack in attacks:
        f1_scores = []
        print(f"Using attack: {attack.__class__.__name__}")

        count = 0
        # 初始化 TP, FN, FP, TN 的数组，用于存储每个阈值下的结果
        TP_array = [0] * len(thresholds)
        FN_array = [0] * len(thresholds)
        FP_array = [0] * len(thresholds)
        TN_array = [0] * len(thresholds)

        # 初始化 precision, recall, accuracy 的数组，用于存储每个阈值下的结果
        precision = []
        recall = []
        accuracy = []

        for i, (X, y) in enumerate(test_loader):
            X, y = X.cuda(), y.cuda()
            # 模拟一个输入
            input_image = X.requires_grad_(True)

            # 获取模型输出
            output = model(input_image)
            loss = output.mean()  # 这里使用均值作为损失，实际使用时可能需要根据任务调整
            loss.backward()

            confidence_array = np.zeros((26, 26))
            confidence_array_ori = np.zeros((26, 26))

            smc_tol=0
            smc_tol_ori=0
            n=0

            # _, _, _, _, perturb_image = yoao(input_image.detach(), y, model, 10,overshoot=0.02, max_iter=1)
            perturb_image = attack(input_image, y)  # 确保输入是正确的形状
            check = model(perturb_image)
            check_label=torch.argmax(check, dim=1).squeeze(0)

            if check_label!=y:

                for i in range(0,26,7):
                    for j in range(0,26,7):

                        perturbed_image_copy= perturb_image.clone()
                        image_copy = input_image.clone()

                        perturbed_image = apply_mask(perturbed_image_copy, (7, 7), (i, j), color=(0, 1, 0))
                        ori_image = apply_mask(image_copy, (7, 7), (i, j), color=(0, 1, 0))
                        output = model(ori_image)
                        output_per = model(perturbed_image)
                        # 计算置信度分数
                        confidence = F.softmax(output, dim=1).max().item()
                        confidence_per = F.softmax(output_per, dim=1).max().item()
                        conf = F.softmax(output, dim=1)
                        conf_per = F.softmax(output_per, dim=1)
                        confidence_array_ori[i,j]=confidence
                        confidence_array[i, j] = confidence_per
                        #  SMC
                        smc = -torch.sum(conf * torch.log2(conf))
                        smc_per = -torch.sum(conf_per * torch.log2(conf_per))
                        smc_tol_ori+=smc
                        smc_tol+=smc_per
                        n=n+1
                        # 可视化
                smc_avr=smc_tol/n
                smc_array_ori_avr=smc_tol_ori/n
                print(smc_array_ori_avr)
                print(smc_avr)
                for k, threshold_value in enumerate(thresholds):
                    # 计算 TP, FN, FP, TN
                    if smc_avr > threshold_value:
                        TP_array[k] = TP_array[k] + 1
                    else:
                        FN_array[k] =  FN_array[k] + 1

                    if smc_array_ori_avr > threshold_value:
                        FP_array[k] =  FP_array[k]  + 1
                    else:
                        TN_array[k] = TN_array[k]+ 1

                count = count + 1
            if count == 100:
                break
                # plt.show()

        precision = []
        recall = []
        accuracy = []
        F1 = []

        for i in range(len(TP_array)):
            if TP_array[i] + FP_array[i] > 0:
                precision.append(TP_array[i] / (TP_array[i] + FP_array[i]))
            else:
                precision.append(0)

            if TP_array[i] + FN_array[i] > 0:
                recall.append(TP_array[i] / (TP_array[i] + FN_array[i]))
            else:
                recall.append(0)

            accuracy.append((TP_array[i] + TN_array[i]) / (TP_array[i] + TN_array[i] + FP_array[i] + FN_array[i]))

            if precision[i] + recall[i] > 0:
                F1.append(2 * (precision[i] * recall[i]) / (precision[i] + recall[i]))
            else:
                F1.append(0)

        # 打印结果
        for i, threshold in enumerate(thresholds):  # thresholds 是阈值数组
            print("threshold {}".format(threshold))
            print("Precision {:.2f}%".format(precision[i] * 100))
            print("Recall {:.2f}%".format(recall[i] * 100))
            print("Accuracy {:.2f}%".format(accuracy[i] * 100))
            print("F1 Score: {:.2f}".format(F1[i] * 100))


            # 将结果写入CSV文件
            csv_file = '3x3_res18_96%_thressholds_metrics-test.xlsx'

            # 打开Excel文件，如果文件不存在则创建
            try:
                wb = openpyxl.load_workbook(csv_file)
            except FileNotFoundError:
                wb = Workbook()

            # 创建新的sheet，名字为Using attack: {attack.__class__.__name__}

            safe_name = re.sub(r'[\\/*?:\[\]<>|]', '', attack.__class__.__name__)
            sheet_name = f"Using attack_{safe_name}"
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
            ws = wb[sheet_name]

            # 写入精准率和召回率等数据
            data = [f"Using attack: {attack.__class__.__name__}", "Precision: {:.2f}%".format(precision[i] * 100),
                    "Recall: {:.2f}%".format(recall[i] * 100), "F1 Score: {:.2f}".format(F1[i] * 100),
                    "accuracy: {:.2f}%".format(accuracy[i] * 100), "thresholds: {:.3f}".format(threshold)]
            ws.append(data)

            # 保存Excel文件
            wb.save(csv_file)


if __name__ == "__main__":
    main()